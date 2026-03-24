import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border
from openpyxl.styles.borders import Side
from datetime import datetime, date
import calendar
import tempfile
import os
from pathlib import Path

director = 'Директор'
start_text_in_a5_ab5 = 'Арендодатель: '
start_text_in_a7_ab7 = 'Арендатор: '
start_text_in_a16_ab16 = 'Всего оказано услуг 1, на сумму '
text_for_user_accept_data_A19_AB19= 'Вышеперечисленные услуги выполнены полностью и в срок. Арендатор претензий по объему, качеству и срокам оказания услуг не имеет.'

def get_text_work_in_act(number_agreement, current_date):
    split_agreement = str(number_agreement).split(' ')
    agr_num = split_agreement[0] if len(split_agreement) > 0 else ""
    agr_date = f" от {split_agreement[1]}" if len(split_agreement) > 1 else ""
    
    start_date = datetime(current_date.year, current_date.month, 1)
    last_day_in_this_month = calendar.monthrange(current_date.year, current_date.month)[1]
    last_date = date(current_date.year, current_date.month, last_day_in_this_month)
    text = f'Услуги аренды по договору аренды нежилого помещения {agr_num}{agr_date}. За период с {start_date.strftime("%d.%m.%Y")} по {last_date.strftime("%d.%m.%Y")}.'
    return text

def sum_propisyu_full(summa):
    units = ["", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
    units_female = ["", "одна", "две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
    teens = ["десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать", 
             "пятнадцать", "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"]
    tens = ["", "десять", "двадцать", "тридцать", "сорок", "пятьдесят", 
            "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
    hundreds = ["", "сто", "двести", "триста", "четыреста", "пятьсот", 
                "шестьсот", "семьсот", "восемьсот", "девятьсот"]
    
    thousands = ["тысяча", "тысячи", "тысяч"]
    millions = ["миллион", "миллиона", "миллионов"]
    
    if isinstance(summa, str):
        # Заменяем запятую на точку, если есть
        summa = summa.replace(',', '.')
        summa = float(summa)
    
    rubli = int(summa)  # int от float работает нормально
    kopeyki = int(round((summa - rubli) * 100))
    
    def number_to_words(num, gender_female=False):
        if num == 0:
            return "ноль"
        
        result = []
        
        # Миллионы
        mil = num // 1_000_000
        if mil > 0:
            mil_text = convert_triple(mil, False)
            result.append(mil_text + " " + get_plural_form(mil, millions))
            num %= 1_000_000
        
        # Тысячи
        thousand = num // 1_000
        if thousand > 0:
            thousand_text = convert_triple(thousand, True)
            result.append(thousand_text + " " + get_plural_form(thousand, thousands))
            num %= 1_000
        
        # Остаток
        if num > 0:
            result.append(convert_triple(num, gender_female))
        
        return " ".join(result)
    
    def convert_triple(num, is_thousand=False):
        """Конвертирует трёхзначное число"""
        if num == 0:
            return ""
        
        res = []
        
        # Сотни
        h = num // 100
        if h > 0:
            res.append(hundreds[h])
            num %= 100
        
        # Десятки и единицы
        if num >= 20:
            d = num // 10
            res.append(tens[d])
            num %= 10
            if num > 0:
                if is_thousand:
                    res.append(units_female[num])
                else:
                    res.append(units[num])
        elif num >= 10:
            res.append(teens[num - 10])
        elif num > 0:
            if is_thousand:
                res.append(units_female[num])
            else:
                res.append(units[num])
        
        return " ".join(res)
    
    def get_plural_form(num, forms):
        """Возвращает правильную форму слова"""
        num = num % 100
        if 11 <= num <= 19:
            return forms[2]
        
        last_digit = num % 10
        if last_digit == 1:
            return forms[0]
        if 2 <= last_digit <= 4:
            return forms[1]
        return forms[2]
    
    def get_rubles_word(num):
        """Склонение слова 'рубль'"""
        num = num % 100
        if 11 <= num <= 19:
            return "рублей"
        
        last_digit = num % 10
        if last_digit == 1:
            return "рубль"
        if 2 <= last_digit <= 4:
            return "рубля"
        return "рублей"
    
    def get_kopeyki_word(num):
        """Склонение слова 'копейка'"""
        num = num % 100
        if 11 <= num <= 19:
            return "копеек"
        
        last_digit = num % 10
        if last_digit == 1:
            return "копейка"
        if 2 <= last_digit <= 4:
            return "копейки"
        return "копеек"
    
    # Получаем рубли прописью
    rubli_text = number_to_words(rubli, False)
    if rubli == 0:
        rubli_text = "ноль"
    else:
        rubli_text = rubli_text.capitalize()
    
    # Формируем результат
    rubli_word = get_rubles_word(rubli)
    kopeyki_word = get_kopeyki_word(kopeyki)
    
    return f"{rubli_text} {rubli_word} {kopeyki:02d} {kopeyki_word}"

def get_short_fio(fio):
    print(fio)
    fio_split = fio.split(' ')
    print(fio_split)
    short_name = f'{fio_split[1][0]}.'.upper()
    short_patronymic = f'{fio_split[2][0]}.'.upper()
    short_fio = f'{fio_split[0]} {short_name}{short_patronymic}'
    return short_fio

async def create_layoat_for_user(name_company, name_company_tenant, current_date, final_price, square, agreement, full_name_tenant, act_number):
    price_rub_cop = str(final_price).split('.')
    short_full_name_tenant = get_short_fio(full_name_tenant)
    thin_top_border  = Border(top=Side(border_style="thin", color="000000"))
    thin_bottom_border  = Border(bottom=Side(border_style="thin", color="000000"))
    if isinstance(current_date, str):
        try:
            # Пробуем разные форматы даты
            for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
                try:
                    current_date = datetime.strptime(current_date, fmt)
                    break
                except ValueError:
                    continue
        except:
            # Если ничего не подошло, используем текущую дату
            current_date = datetime.now()
    wb = load_workbook('docs/create_layout_us.xlsx')
    ws = wb.active
    
    act_text = f'Акт №{str(act_number)} от {current_date.strftime("%d.%m.%Y")}'
    name_file = f'Акт на оплату арендного платежа для {name_company_tenant}'

    ws.merge_cells('A3:AA3')
    ws['A3'] = act_text
    ws['A3'].font = Font(bold=True,size=16)
    for row in ws['A3:AA3']:
        for cell in row:
            cell.border = thin_bottom_border

    ws.merge_cells('A5:AB5')
    ws['A5']= f'{start_text_in_a5_ab5} {name_company}'
    ws['A5'].alignment = Alignment(horizontal='left',vertical='center')
    ws['A5'].font = Font(bold=True)
    
     
    ws.merge_cells('A7:AB7')
    ws['A7'].alignment = Alignment(horizontal='left',vertical='center')
    ws['A7'] = f'{start_text_in_a7_ab7} {name_company_tenant}'
    ws['A7'].font = Font(bold=True)

    text_work_in_act = get_text_work_in_act(agreement, current_date)
    ws.merge_cells('C11:S12')
    ws['C11'] = text_work_in_act
    ws['C11'].alignment = Alignment(horizontal='center',vertical='center')


    ws['T11'] = str(square)
    ws.merge_cells('T11:V12')
    ws['T11'].alignment = Alignment(horizontal='center',vertical='center')

    ws.merge_cells('Y11:Z12')
    ws['Y11'] = final_price
    ws['Y11'].alignment = Alignment(horizontal='center',vertical='center')

    ws.merge_cells('Y14:Z14')
    ws['Y14'] = final_price
    ws['Y14'].font = Font(bold=True,size=14)
    ws['Y14'].alignment = Alignment(horizontal='center',vertical='center')

    ws.merge_cells('A16:AB16')
    ws['A16'] = f'{start_text_in_a16_ab16} {price_rub_cop[0]} руб. {price_rub_cop[1]} коп.'

    ws.merge_cells('A19:AB19')
    ws['A19'] = text_for_user_accept_data_A19_AB19
    ws['A19'].alignment = Alignment(horizontal='left',vertical='center')

    ws.merge_cells('A21:E21')
    ws['A21'] = start_text_in_a5_ab5
    ws['A21'].font = Font(bold=True,size=14)

    ws['T21'] = start_text_in_a7_ab7
    ws.merge_cells('T21:AB21')
    ws['T21'].font = Font(bold=True,size=14)

    ws.merge_cells('A17:AA17')
    ws['A17'] = f'({sum_propisyu_full(final_price)})'

    ws.merge_cells('A22:P22')
    ws['A22'] = f'{director} {name_company}'

    ws.merge_cells('T22:AB22')
    ws['T22'] = f'{director} {name_company_tenant}'

    ws.merge_cells('A24:P24')
    ws['A24'] = 'Попова Я.В.'
    for row in ws['A24:P24']:
        for cell in row:
            cell.border = thin_top_border

    ws.merge_cells('T24:AB24')
    ws['T24'] = short_full_name_tenant
    for row in ws['T24:AB24']:
        for cell in row:
            cell.border = thin_top_border
    
    # Создание временного файла
    temp_dir = tempfile.gettempdir()
    temp_file_path = os.path.join(temp_dir, f"{name_file}_{next(tempfile._get_candidate_names())}.xlsx")
    
    # Сохраняем во временный файл
    wb.save(temp_file_path)
    
    # Возвращаем путь к временному файлу
    return temp_file_path

async def create_invoice_for_payment_for_user(act_number, full_name_company, agreement, price):
    months_ru = {
                1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
                5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
                9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'
            }
    today = datetime.now()
    first_day = today.replace(day=1).strftime("%d.%m.%Y")
    fifth_day = today.replace(day=5).strftime("%d.%m.%Y")
    prev = datetime.now().replace(day=1)
    period_str = f"{months_ru[prev.month]} {prev.year}"
    thin_top_border  = Border(top=Side(border_style="thin", color="000000"))
    thin_bottom_border  = Border(bottom=Side(border_style="thin", color="000000"))
    agreement_list = agreement.split(' ')
    
    wb = load_workbook('docs/invoice_for_payment.xlsx')
    ws = wb.active
    
    act_text = f'Счет на оплату №{str(act_number)} от {first_day}'
    cleaned = full_name_company.replace('"', '')
    name_file = f'Счет на оплату арендного платежа для {cleaned}'

    ws.merge_cells('A5:M5')
    ws['A5'] = act_text
    ws['A5'].font = Font(bold=True,size=16)
    for row in ws['A5:M5']:
        for cell in row:
            cell.border = thin_bottom_border

    ws.merge_cells('A7:M7')
    ws['A7']= f'Покупатель: {full_name_company}'
    ws['A7'].alignment = Alignment(horizontal='left',vertical='center')
    ws['A7'].font = Font(bold=True)
    
     
    ws.merge_cells('C9:D9')
    ws['C9'].alignment = Alignment(horizontal='left',vertical='center')
    
    agr_num = agreement_list[0] if len(agreement_list) > 0 else ""
    agr_date = f" от {agreement_list[1]}" if len(agreement_list) > 1 else ""
    ws['C9'] = f'Услуги аренды за {period_str}. По договору аренды {agr_num}{agr_date}.'
    text_length = len(ws['C9'].value)
    estimated_width = min(text_length * 1.2, 100)  # грубая оценка
    ws.column_dimensions['C'].width = estimated_width / 2
    ws.column_dimensions['D'].width = estimated_width / 2

    ws.merge_cells('C9:D9')

    ws.merge_cells('K9:L9')
    ws['K9'] = price
    ws['K9'].alignment = Alignment(horizontal='center',vertical='center')


    ws.merge_cells('A10:M10')
    ws['A10'] = f'Оплатить до:                                                                   {fifth_day}'
    ws['A10'].alignment = Alignment(horizontal='left',vertical='center')


    ws.merge_cells('L11')
    ws['L11'] = price
    ws['L11'].font = Font(bold=True,size=14)
    ws['L11'].alignment = Alignment(horizontal='center',vertical='center')

    ws.merge_cells('A12:E12')
    ws['A12'] = f'Всего наименований 1, на сумму {price} руб.'

    ws.merge_cells('B13:M13')
    ws['B13'] = f'({sum_propisyu_full(price)})'
    ws['B13'].alignment = Alignment(horizontal='left',vertical='center')

    ws.merge_cells('I9:J9')
    ws['I9'] = price
    

    # script_dir = os.path.dirname(os.path.abspath(__file__))
    # docs_dir = os.path.join(os.path.dirname(script_dir), 'docs')
    # os.makedirs(docs_dir, exist_ok=True)

    # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # file_path = os.path.join(docs_dir, f"invoice_{full_name_company}_{timestamp}.xlsx")

    # wb.save(file_path)
    # print(f"✅ Счёт сохранён: {file_path}")
    # Создание временного файла
    temp_dir = tempfile.gettempdir()
    temp_file_path = os.path.join(temp_dir, f"{name_file}_{next(tempfile._get_candidate_names())}.xlsx")
    
    # Сохраняем во временный файл
    wb.save(temp_file_path)
    
    # Возвращаем путь к временному файлу
    return temp_file_path

