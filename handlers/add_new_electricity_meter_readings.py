from datetime import datetime
import tempfile
from docx import Document
from aiogram import Bot, types, Router
from aiogram.types import Message, CallbackQuery
from aiogram.filters.command import Command
import os
import asyncio
import logging
from aiogram import F
import asyncpg
from handlers.config import config
from aiogram.fsm.state import State
from aiogram.fsm.context import FSMContext
from states.meter_redings_state import Meter_Readings_States
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from states.auth_states import Auth_States

add_new_el_mr_router = Router()

async def get_data(query: str, *params):
    try:
        conn = await asyncpg.connect(config.db_connection)
        result = await conn.fetch(query,*params)
        await conn.close()
        return result
    except Exception as e: 
        print(f"Ошибка: {e}")
        return None
    
async def new_data_insert(query: str, *params):
    try:
        conn = await asyncpg.connect(config.db_connection)
        result = await conn.execute(query,*params)
        return result
    except Exception as e:
        print(f"Ошибка: {e}")
        return None
    
async def get_sheet_name(id_us):
    results = await get_data('SELECT sheets_name FROM users WHERE User_Id = $1',str(id_us))
    if results:
        for list in results:
            sheet_name = list['sheets_name']
    print(sheet_name)
    return sheet_name    

async def add_to_user_list_el(user_id, *values):
    from main import redis as r
    key = f"user:{user_id}:list_el"
    await r.lpush(key, *values) 

def save_keyboard():
    buttons = [
        [types.InlineKeyboardButton(text="Сохранить", callback_data="save_mr_el"), types.InlineKeyboardButton(text="Редактировать", callback_data="redact_mr_el")],
        [types.InlineKeyboardButton(text="Заполнить заново", callback_data="restart_mr_el")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons, resize_keyboard=True)
    return keyboard

def create_main_keyboard():
    buttons = [
        [types.InlineKeyboardButton(text="➕ Добавить ещё", callback_data="add_more_el"), types.InlineKeyboardButton(text="✅ Всё", callback_data="finish_input_el")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons, resize_keyboard=True)
    return keyboard

def generate_keyboard(all_mr, state):
    buttons = []
    if state == 'el':
        for counter, mr in enumerate(all_mr, start=1):
            cb_data = f'mr_el_{mr}'
            str_counter = str(counter)
            buttons.append(
                types.InlineKeyboardButton(text=str_counter, callback_data=cb_data)
            )
        inline_keyboard = []
        for i in range(0, len(buttons), 2):
            inline_keyboard.append(buttons[i:i+2])
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=inline_keyboard)
    return keyboard 

async def check_meter_readings_in_db(id_user):
    records = await get_data('SELECT id_business FROM users WHERE User_Id = $1',str(id_user))
    for record in records:
        id_business = record['id_business']
    result_readings = await get_data('SELECT COUNT(*)<>0 AS check FROM us_readings WHERE business_id = $1',int(id_business))
    for result_reading in result_readings:
        check = result_reading['check']
    return check

async def check_count_meter_readings_in_db(id_user):
    results_user = await get_data('SELECT id FROM users WHERE User_Id = $1', id_user)

@add_new_el_mr_router.message(Meter_Readings_States.add_new_electricity_readings)
async def get_message(msg: Message, state: FSMContext):
    from main import redis as r
    count = 0
    id_us = msg.chat.id
    text_value = (msg.text or "").strip()
    current_state = await state.get_state()
    await msg.delete()
    if text_value.isdigit() and int(text_value) > 0 and len(text_value) < 20:
        number = int(text_value)
        if current_state == Meter_Readings_States.add_new_electricity_readings:
            await state.set_state(Meter_Readings_States.wait_el_mr_state)
            key = f"user:{id_us}:list_el"
        all_mr = await r.lrange(key, 0, -1)
        
        for mr in all_mr:
            if number == int(mr):
                count +=1 
        if count > 0:
            print(all_mr)
            await msg.answer('Вы уже ввели номер этого счетчика',reply_markup=create_main_keyboard())
        else:
            await add_to_user_list_el(id_us, text_value)
            await msg.answer(f'Приняли ваш номер счетчика {text_value}, двигаемся дальше.\nХотите добавить ещё счетчики для ЭЛЕКТРИЧЕСТВА?',reply_markup=create_main_keyboard())
    else:
        await msg.answer('Введите пожалуйста число с счетчика, без других символов')

async def set_new_value(numbers,sheet_name):
    file_path = 'docs/ГИРА_1006теккаа2.xlsx'
    df = pd.read_excel(file_path,sheet_name=sheet_name)
    print(f'Проверка названия листа{sheet_name}')
    parts = []
    text_format = ''
    for number in numbers:
        parts.append(str(number))
    text_format = ','.join(parts)
    column = df.columns[0]
    row_index = df[df[column].astype(str).str.contains('Номер счетчика')].index
    value_position_in_column = 1
    row = row_index[0]
    do_value = df.iloc[row,value_position_in_column]
    if isinstance(do_value, int):
        new_value = f' {text_format}'
    elif isinstance(do_value,bool):
        new_value = f' {text_format}'
    elif isinstance(do_value, float):
        new_value = f' {text_format}'
    else:
        if len(do_value)>0:
            new_value = str(do_value)+ ',' + f' {text_format}'
        else:
            new_value = str(do_value) + f'  {text_format}'
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    excel_row = row + 2
    excel_column = value_position_in_column+1
    ws.cell(row=excel_row, column=excel_column, value=new_value)
    wb.save(file_path)

@add_new_el_mr_router.callback_query(F.data.in_(['add_more_el','finish_input_el', 'save_mr_el', 'redact_mr_el','restart_mr_el']))
async def callback(call: CallbackQuery, state: FSMContext):
    from main import redis as r
    from handlers.reg_user import smart_add_keyboard
    from handlers.run import get_menu_keyboard
    data = call.data
    id_us = call.message.chat.id
    await call.message.delete()
    current_state = await state.get_state()
    if data == 'add_more_el':
        await state.set_state(Meter_Readings_States.add_new_electricity_readings) 
        await call.message.answer('Пожалуйста напишите ещё один номер счетчика ЭЛЕКТРИЧЕСТВА')
    elif data == 'finish_input_el':
        key = f"user:{id_us}:list_el"
        all_mr = await r.lrange(key, 0, -1)
        counter = 1
        all_mr_el_text = 'Проверьте правильность написания\nНомера счетчиков на ЭЛЕКТРИЧЕСТВО:\n'
        for mr in all_mr:
            all_mr_el_text+= f'{str(counter)}) {mr}\n'
            counter+=1
        await call.message.answer(text=all_mr_el_text,reply_markup=save_keyboard())
    elif data == 'restart_mr_el':
        key = f"user:{id_us}:list_el"
        await state.set_state(Meter_Readings_States.add_new_electricity_readings) 
        await r.delete(key)
        await call.message.answer('Пожалуйста введите номер счетчика')
    elif data == 'save_mr_el':
        key = f"user:{id_us}:list_el"
        key2 = f"user:{id_us}:meters"
        id_type = 2
        all_mr = await r.lrange(key, 0, -1)
        records = await get_data('SELECT id_business FROM users WHERE User_Id = $1',str(id_us))
        id_business = records[0]['id_business'] if records else None
        
        for mr in all_mr:
            if id_business:
                await new_data_insert('INSERT INTO us_readings(business_id, number_counter,counter_type_id) VALUES ($1, $2,$3)',int(id_business),mr,id_type)
        sheet_name = await get_sheet_name(id_us)
        await r.delete(key)
        await set_new_value(all_mr,sheet_name)
        # await state.set_state(Auth_States.menu_state)
        not_filled_str = await r.get(key2)
        not_filled = not_filled_str.split(',') if not_filled_str else []
        if 'el' in not_filled:
            not_filled.remove('el')
            meters_str = ','.join(not_filled)
            await r.set(key2, meters_str)
        if len(not_filled)<1:
            await state.set_state(Auth_States.menu_state)
            await call.message.answer('Вы в меню', reply_markup=get_menu_keyboard())
        else:
            keyboard = await smart_add_keyboard(id_us)
            await call.message.answer('Вы в меню, сохранили ваш счетчик', reply_markup=keyboard)
    elif data == 'redact_mr_el':
        if current_state == Meter_Readings_States.add_new_electricity_readings or current_state == Meter_Readings_States.wait_el_mr_state:
            key = f"user:{id_us}:list_el"
            state_wmr = 'el'
        all_mr = await r.lrange(key, 0, -1)
        counter = 1
        all_mr_water_text = 'Пожалуйста нажмите на кнопку, под которой располагается номер вашего неверно введеного счетчика\n'
        print(f'При нажатии кнопки список чек - {all_mr}')
        for mr in all_mr:
            all_mr_water_text+= f'{str(counter)}) {mr}\n'
            counter+=1
        keyboard = generate_keyboard(all_mr,state_wmr)
        await state.set_state(Meter_Readings_States.wait_mr_state_edit)
        await call.message.answer(text = all_mr_water_text,reply_markup=keyboard)

@add_new_el_mr_router.message(Meter_Readings_States.wait_el_mr_state)
async def get_message(msg: Message, state: FSMContext):
    await msg.answer('Пожалуйста нажмите одну из кнопок выше, чтоб мы могли продолжить⬆️')
    
@add_new_el_mr_router.callback_query(F.data.startswith('mr_el_'))
async def cq(call: CallbackQuery, state: FSMContext):
    data = call.data
    number_edit = data[6:]
    await state.update_data(edit_mr_el = number_edit)
    await call.message.delete()
    await state.set_state(Meter_Readings_States.edit_mr_el)
    await call.message.answer('Введите пожалуйста новое значение')

@add_new_el_mr_router.message(Meter_Readings_States.edit_mr_el)
async def msg(msg: Message, state: FSMContext):
    from main import redis as r
    text_value = (msg.text or "").strip()
    if not (text_value.isdigit() and int(text_value) > 0 and len(text_value) < 20):
        await msg.answer('Введите пожалуйста число с счетчика, без других символов')
        return

    new_value = int(text_value)
    id_us = msg.chat.id
    current_state = await state.get_state()
    if current_state == Meter_Readings_States.edit_mr_el: 
        key = f"user:{id_us}:list_el"
    all_inf = await state.get_data()
    editing_value = all_inf['edit_mr_el']
    count = 0
    print(new_value)
    print(editing_value)
    all_mr = await r.lrange(key, 0, -1)
    for mr in all_mr:
        if int(new_value) == int(mr):
            count +=1 
    if count > 0:
        print(all_mr)
        await msg.answer('Вы уже ввели номер этого счетчика, введите повторно')
    else:
        for i, item in enumerate(all_mr):
            if item == editing_value:
                await r.lset(key, i, text_value)
                print(f"✅ Заменили {editing_value} на {text_value} на позиции {i}")
    all_mr = await r.lrange(key, 0, -1)
    counter = 1
    all_mr_water_text = 'Проверьте правильность написания\nНомера счетчиков на ЭЛЕКТРИЧЕСТВО:\n'
    for mr in all_mr:
        all_mr_water_text+= f'{str(counter)}) {mr}\n'
        counter+=1
    await msg.answer(text=all_mr_water_text,reply_markup=save_keyboard())

@add_new_el_mr_router.callback_query(F.data.in_(['selected_mr_el_add']))
async def callback(call: CallbackQuery, state: FSMContext):
    data = call.data
    await call.message.delete()
    if data == 'selected_mr_el_add':
        await state.set_state(Meter_Readings_States.add_new_electricity_readings)
        await call.message.answer('Введите номер счетчика ЭЛЕКТРИЧЕСТВА')