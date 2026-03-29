from datetime import datetime
import tempfile
from html import escape
from docx import Document
from aiogram import Bot, types, Router
from aiogram.types import Message, CallbackQuery
from aiogram.filters.command import Command
import os
import asyncio
import logging
from aiogram import F
import asyncpg
from handlers.config import config
from aiogram.fsm.state import State
from aiogram.fsm.context import FSMContext
from states.meter_redings_state import Meter_Readings_States
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from states.auth_states import Auth_States

add_new_mr_router = Router()

async def get_data(query: str, *params):
    conn = None
    try:
        conn = await asyncpg.connect(config.db_connection)
        return await conn.fetch(query, *params)
    except Exception as e:
        print(f"Ошибка: {e}")
        return None
    finally:
        if conn:
            await conn.close()

async def new_data_insert(query: str, *params):
    conn = None
    try:
        conn = await asyncpg.connect(config.db_connection)
        return await conn.execute(query, *params)
    except Exception as e:
        print(f"Ошибка: {e}")
        return None
    finally:
        if conn:
            await conn.close()

    
async def get_sheet_name(id_us):
    results = await get_data('SELECT sheets_name FROM users WHERE User_Id = $1',str(id_us))
    if results:
        for list in results:
            sheet_name = list['sheets_name']
    print(sheet_name)
    return sheet_name
    
def generate_keyboard(all_mr, state):
    buttons = []
    if state == 'cold':
        for counter, mr in enumerate(all_mr, start=1):
            cb_data = f'mr_cw_{mr}'
            str_counter = str(counter)
            buttons.append(
                types.InlineKeyboardButton(text=str_counter, callback_data=cb_data)
            )
        inline_keyboard = []
        for i in range(0, len(buttons), 2):
            inline_keyboard.append(buttons[i:i+2])
    elif state == 'hot':
        for counter, mr in enumerate(all_mr, start=1):
            cb_data = f'mr_hw_{mr}'
            str_counter = str(counter)
            buttons.append(
                types.InlineKeyboardButton(text=str_counter, callback_data=cb_data)
            )
        inline_keyboard = []
        for i in range(0, len(buttons), 2):
            inline_keyboard.append(buttons[i:i+2])
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=inline_keyboard)
    return keyboard 

def hot_or_cold_keyboard():
    buttons = [
        [types.InlineKeyboardButton(text="Холодная вода", callback_data="selected_mr_cw_add")],
        [types.InlineKeyboardButton(text="Горячая вода", callback_data="selected_mr_hw_add")],
        [types.InlineKeyboardButton(text='В меню', callback_data='add_hoc_go_back_cb')]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons, resize_keyboard=True)
    return keyboard

def save_keyboard():
    buttons = [
        [types.InlineKeyboardButton(text="Сохранить", callback_data="save_mr_w"), types.InlineKeyboardButton(text="Редактировать", callback_data="redact_mr_w")],
        [types.InlineKeyboardButton(text="Заполнить заново", callback_data="restart_mr_w")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons, resize_keyboard=True)
    return keyboard

def create_main_keyboard():
    buttons = [
        [types.InlineKeyboardButton(text="➕ Добавить ещё", callback_data="add_more"), types.InlineKeyboardButton(text="✅ Всё", callback_data="finish_input")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons, resize_keyboard=True)
    return keyboard
    
async def add_to_user_list_cw(user_id, *values):
    from main import redis as r
    key = f"user:{user_id}:list_cold_water"
    await r.lpush(key, *values)    

async def add_to_user_list_hw(user_id, *values):
    from main import redis as r
    key = f"user:{user_id}:list_hot_water"
    await r.lpush(key, *values)    

async def check_meter_readings_in_db(id_user):
    records = await get_data('SELECT id_business FROM users WHERE User_Id = $1',str(id_user))
    for record in records:
        id_business = record['id_business']
    result_readings = await get_data('SELECT COUNT(*)<>0 AS check FROM us_readings WHERE business_id = $1',int(id_business))
    for result_reading in result_readings:
        check = result_reading['check']
    return check

async def check_count_meter_readings_in_db(id_user):
    results_user = await get_data('SELECT id FROM users WHERE User_Id = $1', id_user)

@add_new_mr_router.message(Meter_Readings_States.add_new_hot_water_readings)
@add_new_mr_router.message(Meter_Readings_States.add_new_cold_water_readings)
async def get_message(msg: Message, state: FSMContext):
    from main import redis as r
    count = 0
    id_us = msg.chat.id
    text_value = (msg.text or "").strip()
    current_state = await state.get_state()
    await msg.delete()
    if text_value.isdigit() and int(text_value) > 0 and len(text_value) < 20:
        number = int(text_value)
        if current_state == Meter_Readings_States.add_new_hot_water_readings:
            await state.set_state(Meter_Readings_States.wait_hw_mr_state)
            key = f"user:{id_us}:list_hot_water"
        elif current_state == Meter_Readings_States.add_new_cold_water_readings:
            await state.set_state(Meter_Readings_States.wait_cw_mr_state)
            key = f"user:{id_us}:list_cold_water"
        all_mr = await r.lrange(key, 0, -1)
        
        for mr in all_mr:
            if number == int(mr):
                count +=1 
        if count >0:
            print(all_mr)
            await msg.answer('Вы уже ввели номер этого счетчика',reply_markup=create_main_keyboard())
        else:
            if current_state == Meter_Readings_States.add_new_hot_water_readings:
                await add_to_user_list_hw(id_us, text_value)
                await msg.answer(f'Приняли ваш номер счетчика {text_value}, двигаемся дальше.\nХотите добавить ещё счетчики для ВОДЫ?',reply_markup=create_main_keyboard())

            elif current_state == Meter_Readings_States.add_new_cold_water_readings:
                await add_to_user_list_cw(id_us, text_value)
                await msg.answer('Приняли ваш номер счетчика, двигаемся дальше.\nХотите добавить ещё счетчики для ВОДЫ?',reply_markup=create_main_keyboard())
    else:
        await msg.answer('Введите пожалуйста число с счетчика, без других символов')

async def set_new_value(numbers,sheet_name):
    file_path = 'docs/ГИРА_1006теккаа2.xlsx'
    df = pd.read_excel(file_path,sheet_name=sheet_name)
    print(f'Проверка названия листа{sheet_name}')
    parts = []
    text_format = ''
    for number in numbers:
        parts.append(str(number))
    text_format = ','.join(parts)
    column = df.columns[0]
    row_index = df[df[column].astype(str).str.contains('Номер счетчика')].index
    value_position_in_column = 1
    row = row_index[0]
    do_value = df.iloc[row,value_position_in_column]
    if isinstance(do_value, int):
        new_value = f' {text_format}'
    elif isinstance(do_value,bool):
        new_value = f' {text_format}'
    elif isinstance(do_value, float):
        new_value = f' {text_format}'
    else:
        if len(do_value)>0:
            new_value = str(do_value)+ ',' + f' {text_format}'
        else:
            new_value = str(do_value) + f'  {text_format}'
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    excel_row = row + 2
    excel_column = value_position_in_column+1
    ws.cell(row=excel_row, column=excel_column, value=new_value)
    wb.save(file_path)

@add_new_mr_router.callback_query(F.data.in_(['selected_mr_cw_add','selected_mr_hw_add','add_hoc_go_back_cb']))
async def callback(call: CallbackQuery, state: FSMContext):
    from handlers.run import build_menu_keyboard
    data = call.data
    await call.message.delete()
    if data == 'selected_mr_cw_add':
        await state.set_state(Meter_Readings_States.add_new_cold_water_readings)
        await call.message.answer('Введите номер счетчика ХОЛОДНОЙ воды')
    elif data == 'selected_mr_hw_add':
        await state.set_state(Meter_Readings_States.add_new_hot_water_readings)
        await call.message.answer('Введите номер счетчика ГОРЯЧЕЙ воды')
    elif data == 'add_hoc_go_back_cb':
        await state.set_state(Auth_States.menu_state)
        await call.message.answer('Вы в меню', reply_markup=await build_menu_keyboard(call.from_user.id))

@add_new_mr_router.callback_query(F.data.in_(['add_more','finish_input', 'save_mr_w', 'redact_mr_w','restart_mr_w']))
async def callback(call: CallbackQuery, state: FSMContext):
    from main import redis as r
    from handlers.run import build_menu_keyboard
    from handlers.reg_user import smart_add_keyboard
    data = call.data
    id_us = call.message.chat.id
    await call.message.delete()
    current_state = await state.get_state()
    if data == 'add_more':
        if current_state == Meter_Readings_States.wait_hw_mr_state:
            await state.set_state(Meter_Readings_States.add_new_hot_water_readings) 
        elif current_state == Meter_Readings_States.wait_cw_mr_state:
            await state.set_state(Meter_Readings_States.add_new_cold_water_readings) 
        await call.message.answer('Пожалуйста напишите ещё один номер счетчика воды')
    elif data == 'finish_input':
        if current_state == Meter_Readings_States.add_new_hot_water_readings or current_state == Meter_Readings_States.wait_hw_mr_state:
            key = f"user:{id_us}:list_hot_water"
        elif current_state == Meter_Readings_States.add_new_cold_water_readings or current_state == Meter_Readings_States.wait_cw_mr_state:
            key = f"user:{id_us}:list_cold_water"
        all_mr = await r.lrange(key, 0, -1)
        counter = 1
        all_mr_water_text = 'Проверьте правильность написания\nНомера счетчиков на воду:\n'
        for mr in all_mr:
            all_mr_water_text+= f'{str(counter)}) {mr}\n'
            counter+=1
        await call.message.answer(text=all_mr_water_text,reply_markup=save_keyboard())
    elif data == 'restart_mr_w':
        if current_state == Meter_Readings_States.add_new_hot_water_readings or current_state == Meter_Readings_States.wait_hw_mr_state:
            key = f"user:{id_us}:list_hot_water"
            await state.set_state(Meter_Readings_States.add_new_hot_water_readings) 
        elif current_state == Meter_Readings_States.add_new_cold_water_readings or current_state == Meter_Readings_States.wait_cw_mr_state:
            key = f"user:{id_us}:list_cold_water"
            await state.set_state(Meter_Readings_States.add_new_cold_water_readings) 
        await r.delete(key)
        
        await call.message.answer('Пожалуйста введите номер счетчика')
    elif data == 'save_mr_w':
        key2 = f"user:{id_us}:meters"
        not_filled_str = await r.get(key2)
        not_filled = not_filled_str.split(',') if not_filled_str else []
        if current_state == Meter_Readings_States.add_new_hot_water_readings or current_state == Meter_Readings_States.wait_hw_mr_state:
            key = f"user:{id_us}:list_hot_water"
            if 'hw' in not_filled:
                not_filled.remove('hw')
                meters_str = ','.join(not_filled)
                await r.set(key2, meters_str)
            id_type = 3
        elif current_state == Meter_Readings_States.add_new_cold_water_readings or current_state == Meter_Readings_States.wait_cw_mr_state:
            key = f"user:{id_us}:list_cold_water"
            id_type = 1
            if 'cw' in not_filled:
                not_filled.remove('cw')
                meters_str = ','.join(not_filled)
                await r.set(key2, meters_str)
        all_mr = await r.lrange(key, 0, -1)
        records = await get_data('SELECT id_business FROM users WHERE User_Id = $1',str(id_us))
        id_business = records[0]['id_business'] if records else None
        
        for mr in all_mr:
            if id_business:
                await new_data_insert('INSERT INTO us_readings(business_id, number_counter,counter_type_id) VALUES ($1, $2,$3)',int(id_business),mr,id_type)
        sheet_name = await get_sheet_name(id_us)
        await r.delete(key)
        if len(not_filled)<1:
            try:
                await set_new_value(all_mr,sheet_name)
                await state.set_state(Auth_States.menu_state)
                await call.message.answer('Вы в меню', reply_markup=await build_menu_keyboard(id_us))
            except Exception as e:
                await state.set_state(Auth_States.menu_state)
                await call.message.answer('Вы в меню', reply_markup=await build_menu_keyboard(id_us))
        else:
            keyboard = await smart_add_keyboard(id_us)
            await call.message.answer('Вы в меню, сохранили ваш счетчик', reply_markup=keyboard)
    elif data == 'redact_mr_w':
        if current_state == Meter_Readings_States.add_new_hot_water_readings or current_state == Meter_Readings_States.wait_hw_mr_state:
            key = f"user:{id_us}:list_hot_water"
            state_wmr = 'hot'
        elif current_state == Meter_Readings_States.add_new_cold_water_readings or current_state == Meter_Readings_States.wait_cw_mr_state:
            key = f"user:{id_us}:list_cold_water"
            state_wmr = 'cold'
        all_mr = await r.lrange(key, 0, -1)
        counter = 1
        all_mr_water_text = 'Пожалуйста нажмите на кнопку, под которой располагается номер вашего неверно введеного счетчика\n'
        print(f'При нажатии кнопки список чек - {all_mr}')
        for mr in all_mr:
            all_mr_water_text+= f'{str(counter)}) {mr}\n'
            counter+=1
        keyboard = generate_keyboard(all_mr,state_wmr)
        await state.set_state(Meter_Readings_States.wait_mr_state_edit)
        await call.message.answer(text = all_mr_water_text,reply_markup=keyboard)

@add_new_mr_router.message(Meter_Readings_States.wait_cw_mr_state)
async def get_message(msg: Message, state: FSMContext):
    await msg.answer('Пожалуйста нажмите одну из кнопок выше, чтоб мы могли продолжить⬆️')

@add_new_mr_router.message(Meter_Readings_States.wait_hw_mr_state)
async def get_message(msg: Message, state: FSMContext):
    await msg.answer('Пожалуйста нажмите одну из кнопок выше, чтоб мы могли продолжить⬆️')
    
@add_new_mr_router.callback_query(F.data.startswith('mr_cw_'))
async def cq(call: CallbackQuery, state: FSMContext):
    data = call.data
    number_edit = '_'.join(data.split('_')[2:])
    await state.update_data(edit_mr_w = number_edit)
    await call.message.delete()
    await state.set_state(Meter_Readings_States.edit_mr_cw)
    await call.message.answer(
        f'Введите новое значение для счётчика холодной воды № <b>{escape(str(number_edit))}</b>',
        parse_mode='HTML',
    )

@add_new_mr_router.callback_query(F.data.startswith('mr_hw_'))
async def cq(call: CallbackQuery, state: FSMContext):
    data = call.data
    number_edit = '_'.join(data.split('_')[2:])
    await state.update_data(edit_mr_w = number_edit)
    await call.message.delete()
    await state.set_state(Meter_Readings_States.edit_mr_hw)
    await call.message.answer(
        f'Введите новое значение для счётчика горячей воды № <b>{escape(str(number_edit))}</b>',
        parse_mode='HTML',
    )

@add_new_mr_router.message(Meter_Readings_States.edit_mr_hw)
@add_new_mr_router.message(Meter_Readings_States.edit_mr_cw)
async def msg(msg: Message, state: FSMContext):
    from main import redis as r
    text_value = (msg.text or "").strip()
    if not (text_value.isdigit() and int(text_value) > 0 and len(text_value) < 20):
        editing = (await state.get_data()).get('edit_mr_w', '')
        hint = f'Счётчик №<b>{escape(str(editing))}</b>. ' if editing else ''
        await msg.answer(
            f'{hint}Введите пожалуйста число с счётчика, без других символов',
            parse_mode='HTML',
        )
        return

    new_value = int(text_value)
    id_us = msg.chat.id
    current_state = await state.get_state()
    if current_state == Meter_Readings_States.edit_mr_hw: 
        key = f"user:{id_us}:list_hot_water"
    elif current_state == Meter_Readings_States.edit_mr_cw:
        key = f"user:{id_us}:list_cold_water"
    all_inf = await state.get_data()
    editing_value = all_inf['edit_mr_w']
    count = 0
    print(new_value)
    print(editing_value)
    if isinstance(new_value, int) and int(new_value)>0 and len(text_value)<20:
        print('мы тут')
        all_mr = await r.lrange(key, 0, -1)
        for mr in all_mr:
            if int(new_value) == int(mr):
                count +=1 
        if count > 0:
            print(all_mr)
            await msg.answer('Вы уже ввели номер этого счетчика, введите повторно')
        else:
            for i, item in enumerate(all_mr):
                if item == editing_value:
                    await r.lset(key, i, text_value)
                    print(f"✅ Заменили {editing_value} на {text_value} на позиции {i}")
        all_mr = await r.lrange(key, 0, -1)
        counter = 1
        all_mr_water_text = 'Проверьте правильность написания\nНомера счетчиков на воду:\n'
        for mr in all_mr:
            all_mr_water_text+= f'{str(counter)}) {mr}\n'
            counter+=1
        await msg.answer(text=all_mr_water_text,reply_markup=save_keyboard())
    else:
        hint = f'Счётчик №<b>{escape(str(editing_value))}</b>. ' if editing_value else ''
        await msg.answer(
            f'{hint}Введите пожалуйста число с счётчика, без других символов',
            parse_mode='HTML',
        )

